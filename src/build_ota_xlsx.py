#!/usr/bin/env python3
"""Build the Xuru Stays OTA listing pack.

Follows the house format (Almaza/Kennah packs): title in A1, italic explainer in A2,
headers on row 4, freeze at A5, autofilter, not-eligible rows red-filled and sorted
to the bottom, green->red heatmap on the monthly price columns.

Adds the Zen Stays "Send Back" tab, because Xuru's public catalogue has real gaps
(missing amenities, empty galleries, unknown availability) and nothing here is
guessed or auto-corrected — the gaps go back to the operator as questions.
"""
import json, glob, os, re, collections, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# The repo's own catalogue, refreshed by catalogue.js on every CI run. It used to
# default to a session scratchpad, which silently emptied and produced a workbook
# with zero units — see the guard below.
CAT = os.environ.get('XURU_CATALOGUE', os.path.join(ROOT, 'data', 'units'))
TOTALS = os.path.join(ROOT, 'data', 'month-totals.json')
NIGHTS = os.path.join(ROOT, 'data', 'night-rates.json')
ICAL_BASE = 'https://mohamedmaged3002-droid.github.io/xuru-tracker'
SITE_BASE = 'https://bluekeys.co/listings'
OUT = '/Users/MAGED/inv/Xuru OTA Listing Pack.xlsx'
FX = 50
TODAY = dt.date(2026, 8, 13)

HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(color='FFFFFF', bold=True)
RED      = PatternFill('solid', fgColor='FFC7CE')
WARN     = PatternFill('solid', fgColor='FCE4D6')

# --- load ------------------------------------------------------------------
units = []
for f in glob.glob(os.path.join(CAT, '*.json')):
    d = json.load(open(f)); d = d.get('data', d)
    if isinstance(d, dict) and d.get('id'):
        units.append(d)
units.sort(key=lambda u: (u['city'], u['id']))
if len(units) < 100:
    raise SystemExit(f"REFUSING TO BUILD: only {len(units)} units found in {CAT}. "
                     "Expected ~166 — run `node src/catalogue.js` first. A workbook built "
                     "from a partial catalogue looks complete and is not.")
totals = json.load(open(TOTALS)) if os.path.exists(TOTALS) else {}

# Per-night rates are the real source now. Month columns are derived from them:
# median of the actual nightly rates in that month, which is what an OTA rate
# plan wants, rather than a whole-month quote divided by its nights.
nights = json.load(open(NIGHTS))['units'] if os.path.exists(NIGHTS) else {}
night_by_xuru, night_month = {}, {}
for wp, u in nights.items():
    xid = u.get('xuru_id')
    if xid is None: continue
    night_by_xuru[xid] = u
    per = collections.defaultdict(list)
    for day, usd in (u.get('rates') or {}).items():
        per[day[:7]].append(usd)
    night_month[xid] = {m: sorted(v)[len(v)//2] for m, v in per.items() if v}

# wp_post_id + source_code assignment must match the DB insert exactly.
WP = {u['id']: 93001 + i for i, u in enumerate(units)}
CODE = {u['id']: 'XU%03d' % (i + 1) for i, u in enumerate(units)}

def _slugify(t):
    t = re.sub(r'[^a-zA-Z0-9]+', '-', (t or '')).strip('-').lower()
    return re.sub(r'-+', '-', t)[:70].strip('-')
# must match units.slug exactly — the bluekeys_url column is only useful if it resolves
SLUG = {u['id']: f"xuru-{_slugify(u.get('headline'))}-{u['id']}" for u in units}

# Same mapping table as the DB insert (kept in sync by hand — see build_xuru_sql.py).
RULES = [
    (r'mangroovy','Mangroovy Residence','El Gouna','El Gouna'),
    (r'\bg[- ]?cribs\b','G-Cribs','El Gouna','El Gouna'),
    (r'joubal','Joubal','El Gouna','El Gouna'),
    (r'ancient hills','Ancient Hills','El Gouna','El Gouna'),
    (r'ancient sands|ancient s','Ancient Sands','El Gouna','El Gouna'),
    (r'tawila','Tawila','El Gouna','El Gouna'),
    (r'fanadir','Fanadir','El Gouna','El Gouna'),
    (r'maraya','Maraya','El Gouna','El Gouna'),
    (r'sabina','Sabina','El Gouna','El Gouna'),
    (r'cyan','Cyan Residence','El Gouna','El Gouna'),
    (r'\bscarab\b','Scarab Club','El Gouna','El Gouna'),
    (r'\bdawar\b|el gouna|gouna','El Gouna','El Gouna','El Gouna'),
    (r'nazlit|nazlet|al haram|el haram|king faisal|pyramid','Pyramids District','Pyramids District','Giza'),
    (r'mohandessin|dokki|agouza','Mohandessin','Giza','Giza'),
    (r'valory|valore|misr taameer|moltaqa el araby','Valory Sheraton','Heliopolis','Cairo'),
    (r'sheraton al matar|ashbelya|abdelhamid badawy|el nozha|nozha|sheraton','Sheraton Heliopolis','Heliopolis','Cairo'),
    (r'heliopolis|ankara st|ocean blue','Heliopolis','Heliopolis','Cairo'),
    (r'bab al louq|abdeen|bustan|downtown|tahrir|qasr el nil|dar el shefa','Downtown Cairo','Downtown Cairo','Cairo'),
    (r'zamalek','Zamalek','Zamalek','Cairo'),
    (r'maadi','Maadi','Maadi','Cairo'),
    (r'new cairo|fifth settlement|tagamoa','New Cairo','New Cairo','Cairo'),
    (r'nasr city','Nasr City','Nasr City','Cairo'),
]
def classify(u):
    blob = ' '.join([u.get('location_address1') or '', u.get('location_address2') or '',
                     u.get('headline') or '']).lower()
    for rx, comp, area, city in RULES:
        if re.search(rx, blob): return comp, area, city
    return '', '', u['city']

MONTHS = sorted({m for v in night_month.values() for m in v})[:13] or \
         sorted({k for t in totals.values() for k in t.get('months', {})})[:13]

def month_label(key):
    y, m = key.split('-')
    return dt.date(int(y), int(m), 1).strftime("%b '%y")

def style_header(ws, headers, title, explainer):
    ws['A1'] = title; ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = explainer; ws['A2'].font = Font(italic=True, color='808080', size=10)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

wb = Workbook()

# ---------------------------------------------------------------- Tab 1: Master
ws = wb.active; ws.title = 'Xuru Master'
H = ['wp_post_id','code','operator_unit_code','title','city','area','compound','type',
     'bedrooms','baths','guests (BK)','guests (Xuru)','nightly USD (from)','cleaning USD',
     'deposit USD','photos','amenities','availability','blocked nights (365)',
     'lat','lng','ical_url','bluekeys_url','xuru_url','priced nights','ota_eligible','why_not']
style_header(ws, H, 'Xuru Stays — OTA Listing Pack',
    f'Generated {TODAY}. 166 units staged as BlueKeys drafts (wp 93001–93166). '
    f'Prices are Xuru\'s own USD rates — BlueKeys takes 0% (partner, nightly-only). '
    f'"guests (BK)" is bedrooms × 2 per house rule; Xuru\'s own figure is shown beside it.')

rows = []
for u in units:
    comp, area, city = classify(u)
    t = totals.get(str(u['id']), {})
    nphoto = len(u.get('images') or [])
    namen = len(u.get('amenities') or [])
    availability = t.get('availability') or ('ok' if (u.get('availabilities') or {}).get('map') else 'UNKNOWN')
    reasons = []
    if availability != 'ok': reasons.append('no availability data from Xuru')
    if nphoto == 0:  reasons.append('no photos')
    elif nphoto < 8: reasons.append(f'only {nphoto} photos (OTAs want 10+)')
    priced = len((night_by_xuru.get(u['id']) or {}).get('rates') or {})
    if priced == 0:
        reasons.append('no prices yet — renders BLOCKED until the next sweep')
    eligible = 'YES' if not reasons else 'NO'
    beds = int(u.get('number_of_bedrooms') or 0)
    rows.append({
        'sort': (eligible == 'NO', u['city'], u['id']),
        'vals': [WP[u['id']], CODE[u['id']], u.get('name'), u.get('headline'), city, area, comp,
                 (u.get('rental_type') or '').title(), beds, u.get('number_of_bathrooms'),
                 max(beds,1)*2, u.get('capacity_of_guests'),
                 float(u.get('price_value') or 0), float(u.get('cleaning_fee') or 0),
                 u.get('damage_deposit'), nphoto, namen, availability,
                 t.get('blocked_365'), None if not u.get('location_lat') else float(u['location_lat']),
                 None if not u.get('location_lng') else float(u['location_lng']),
                 f"{ICAL_BASE}/{WP[u['id']]}.ics",
                 f"{SITE_BASE}/{SLUG[u['id']]}",
                 f"https://www.xurustays.com/unit-details/{u['id']}",
                 priced, eligible, '; '.join(reasons)],
        'eligible': eligible})
rows.sort(key=lambda r: r['sort'])
for i, r in enumerate(rows, 5):
    for c, v in enumerate(r['vals'], 1):
        cell = ws.cell(row=i, column=c, value=v)
        if r['eligible'] == 'NO': cell.fill = RED
for col, w in zip('ABCDEFGHIJKLMNOPQRSTUVWXYZAA',
                  [11,8,20,46,10,18,20,11,10,7,11,12,15,13,13,8,10,13,16,11,11,62,52,44,13,12,40]):
    ws.column_dimensions[col].width = w

# -------------------------------------------------------- Tab 2: Monthly Prices
ws2 = wb.create_sheet('Monthly Prices')
H2 = ['wp_post_id','code','title','area','beds'] + [month_label(m) for m in MONTHS]
style_header(ws2, H2, 'Xuru Stays — average nightly rate by month (USD)',
    'MEDIAN of the real per-night rates we hold for that month — every night is priced '
    'individually, so this is a true midpoint, not a monthly average. Blocked nights are '
    'excluded (they are never priced). Blank = no rates held for that month yet.')
for i, u in enumerate(units, 5):
    comp, area, city = classify(u)
    nm = night_month.get(u['id'], {})
    vals = [WP[u['id']], CODE[u['id']], u.get('headline'), area, u.get('number_of_bedrooms')]
    vals += [nm.get(m) for m in MONTHS]
    for c, v in enumerate(vals, 1):
        ws2.cell(row=i, column=c, value=v)
if MONTHS:
    first = get_column_letter(6); last = get_column_letter(5 + len(MONTHS))
    ws2.conditional_formatting.add(f'{first}5:{last}{4+len(units)}',
        ColorScaleRule(start_type='min', start_color='C6EFCE',
                       mid_type='percentile', mid_value=50, mid_color='FFEB84',
                       end_type='max', end_color='F8696B'))
for col, w in zip('ABCDE', [11,8,46,18,7]): ws2.column_dimensions[col].width = w
for c in range(6, 6 + len(MONTHS)): ws2.column_dimensions[get_column_letter(c)].width = 10

# --------------------------------------------------- Tab 3: Send Back to Xuru
ws3 = wb.create_sheet('Send Back to Xuru')
H3 = ['#','Units','Issue','Why it matters','What we need']
style_header(ws3, H3, 'Xuru Stays — open questions',
    'Gaps found in Xuru\'s public catalogue. Nothing here has been guessed or auto-corrected.')

no_amen  = [u for u in units if not (u.get('amenities') or [])]
no_photo = [u for u in units if not (u.get('images') or [])]
thin     = [u for u in units if 0 < len(u.get('images') or []) < 8]
no_avail = [u for u in units if not (u.get('availabilities') or {}).get('map')]
no_summ  = [u for u in units if not u.get('summary')]
# Their API cuts `summary` at ~150 chars mid-word, and the full sentence exists
# nowhere else in the payload — so this is theirs to fix, not ours to guess.
trunc_summary = [u for u in units
                 if (u.get('summary') or '').rstrip().endswith('...')
                 and len(u.get('summary') or '') >= 140]
def ids(us, n=6):
    s = ', '.join(f"{CODE[u['id']]} ({u['id']})" for u in us[:n])
    return s + (f' … +{len(us)-n} more' if len(us) > n else '')

issues = [
 (f'{len(no_amen)} units', 'No amenity list published',
  'Amenities drive OTA search filters and cannot be inferred from prose without inventing facts.',
  f'An amenity export for these units. {ids(no_amen)}'),
 (f'{len(no_photo)} units', 'Zero photos in the catalogue API',
  'Cannot be listed anywhere; Booking.com requires 10+ and rejects low-resolution sets.',
  f'Original full-resolution photos (not web renditions). {ids(no_photo)}'),
 (f'{len(thin)} units', 'Fewer than 8 photos',
  'Below the practical OTA threshold; conversion suffers badly.',
  f'More photos. {ids(thin)}'),
 (f'{len(no_avail)} units', 'Availability calendar is empty',
  'An empty calendar means "unknown", never "free" — we will not advertise these as bookable.',
  f'Confirm whether these are live, and re-sync their calendar. {ids(no_avail)}'),
 (f'{len(no_summ)} units', 'No summary text',
  'Thin listing copy; we write our own but need the facts to be right.',
  f'A one-line description. {ids(no_summ)}'),
 ('all 166', 'Per-date rates are only reachable by quoting one night at a time',
  'We currently probe your booking API ~2,000 times a night to track rates. Read-only '
  'BookingSync API access would replace that with one call per unit and remove the load '
  'from your site entirely.',
  'A read-only BookingSync API key, or a scheduled rate + min-stay export.'),
 (f'{len(trunc_summary)} units', 'Listing summary is cut off mid-word in your API',
  'The summary field truncates at ~150 characters ("…dining options are at w..."), and the '
  'full sentence is not recoverable from any other field, so the listing opens mid-thought.',
  f'The untruncated summary text. {ids(trunc_summary)}'),
 ('all 166', 'Minimum-stay rules are not exposed',
  'We may quote a stay Xuru would reject, which surfaces as a failed booking to the guest.',
  'Per-unit (ideally per-date) minimum-stay rules.'),
]
issues = [r for r in issues if not r[0].startswith('0 ')]   # drop resolved issues
for i, row in enumerate(issues, 5):
    ws3.cell(row=i, column=1, value=i - 4)
    for c, v in enumerate(row, 2): ws3.cell(row=i, column=c, value=v)
    for c in range(1, 6): ws3.cell(row=i, column=c).alignment = Alignment(wrap_text=True, vertical='top')
    ws3.row_dimensions[i].height = 46
for col, w in zip('ABCDE', [5,14,34,52,60]): ws3.column_dimensions[col].width = w

wb.save(OUT)
print('wrote', OUT)
print(f'  units {len(units)} | eligible {sum(1 for r in rows if r["eligible"]=="YES")} '
      f'| not eligible {sum(1 for r in rows if r["eligible"]=="NO")} | months {len(MONTHS)}')
